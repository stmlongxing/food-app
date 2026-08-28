import os
import io
import re
import sqlite3
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

try:
    import psycopg2
    from psycopg2.extras import RealDictCursor
except ImportError:
    psycopg2 = None

try:
    from google import genai
    from google.genai import types
except ImportError:
    genai = None

app = Flask(__name__)
DB_NAME = 'inventory.db'
DATABASE_URL = os.environ.get("DATABASE_URL", "")

def format_date_str(val):
    if not val:
        return ""
    val_str = str(val).strip()
    if 'T' in val_str:
        return val_str.split('T')[0]
    if len(val_str) >= 10 and val_str[4] == '-' and val_str[7] == '-':
        return val_str[:10]
    return val_str

class DBConn:
    def __init__(self):
        self.is_pg = bool(DATABASE_URL and psycopg2)
        if self.is_pg:
            url = DATABASE_URL
            if url.startswith("postgres://"):
                url = url.replace("postgres://", "postgresql://", 1)
            self.conn = psycopg2.connect(url, cursor_factory=RealDictCursor)
        else:
            self.conn = sqlite3.connect(DB_NAME)
            self.conn.row_factory = sqlite3.Row

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if exc_type is None:
            self.conn.commit()
        else:
            self.conn.rollback()
        self.conn.close()

    def execute(self, sql, params=()):
        cursor = self.conn.cursor()
        if self.is_pg:
            sql = sql.replace('?', '%s')
        cursor.execute(sql, params)
        return cursor

def init_db():
    with DBConn() as db:
        if db.is_pg:
            db.execute("""
                CREATE TABLE IF NOT EXISTS inventory (
                    id SERIAL PRIMARY KEY,
                    name TEXT NOT NULL,
                    initial_qty INTEGER DEFAULT 0,
                    current_qty INTEGER DEFAULT 0,
                    unit TEXT DEFAULT '罐',
                    mfg_date TEXT,
                    exp_date TEXT,
                    in_date TEXT,
                    staff TEXT,
                    last_audit_date TEXT,
                    auditor TEXT,
                    notes TEXT,
                    status TEXT DEFAULT 'active',
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                );
            """)
            db.execute("""
                CREATE TABLE IF NOT EXISTS audit_logs (
                    id SERIAL PRIMARY KEY,
                    item_id INTEGER,
                    name TEXT,
                    initial_qty INTEGER,
                    in_date TEXT,
                    audit_date TEXT,
                    actual_qty INTEGER,
                    auditor TEXT,
                    notes TEXT,
                    action_type TEXT DEFAULT 'audit',
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                );
            """)
        else:
            db.execute("""
                CREATE TABLE IF NOT EXISTS inventory (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL,
                    initial_qty INTEGER DEFAULT 0,
                    current_qty INTEGER DEFAULT 0,
                    unit TEXT DEFAULT '罐',
                    mfg_date TEXT,
                    exp_date TEXT,
                    in_date TEXT,
                    staff TEXT,
                    last_audit_date TEXT,
                    auditor TEXT,
                    notes TEXT,
                    status TEXT DEFAULT 'active',
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                );
            """)
            db.execute("""
                CREATE TABLE IF NOT EXISTS audit_logs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    item_id INTEGER,
                    name TEXT,
                    initial_qty INTEGER,
                    in_date TEXT,
                    audit_date TEXT,
                    actual_qty INTEGER,
                    auditor TEXT,
                    notes TEXT,
                    action_type TEXT DEFAULT 'audit',
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                );
            """)

try:
    init_db()
except Exception as e:
    print("Init DB error:", e)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/items', methods=['GET'])
def get_items():
    with DBConn() as db:
        cursor = db.execute("SELECT * FROM inventory ORDER BY id DESC")
        items = cursor.fetchall()
    
    result = []
    for item in items:
        d = dict(item)
        d['mfg_date'] = format_date_str(d.get('mfg_date'))
        d['exp_date'] = format_date_str(d.get('exp_date'))
        d['in_date'] = format_date_str(d.get('in_date'))
        d['last_audit_date'] = format_date_str(d.get('last_audit_date'))
        result.append(d)
    return jsonify(result)

@app.route('/api/logs', methods=['GET'])
def get_logs():
    with DBConn() as db:
        cursor = db.execute("SELECT * FROM audit_logs ORDER BY id DESC")
        logs = cursor.fetchall()
    
    result = []
    for l in logs:
        d = dict(l)
        d['in_date'] = format_date_str(d.get('in_date'))
        d['audit_date'] = format_date_str(d.get('audit_date'))
        result.append(d)
    return jsonify(result)

@app.route('/api/logs/<int:log_id>', methods=['PUT'])
def update_log(log_id):
    """修改單筆盤點紀錄"""
    data = request.json or {}
    actual_qty = int(data.get('actual_qty', 0))
    audit_date = format_date_str(data.get('audit_date', ''))
    auditor = data.get('auditor', '').strip()
    notes = data.get('notes', '').strip()
    action_type = 'completed' if actual_qty == 0 else 'audit'

    with DBConn() as db:
        db.execute("""
            UPDATE audit_logs
            SET actual_qty = ?, audit_date = ?, auditor = ?, notes = ?, action_type = ?
            WHERE id = ?
        """, (actual_qty, audit_date, auditor, notes, action_type, log_id))
    return jsonify({"success": True})

@app.route('/api/logs/<int:log_id>', methods=['DELETE'])
def delete_log(log_id):
    """刪除單筆歷史紀錄"""
    with DBConn() as db:
        db.execute("DELETE FROM audit_logs WHERE id = ?", (log_id,))
    return jsonify({"success": True})

@app.route('/api/items', methods=['POST'])
def add_item():
    data = request.json or {}
    name = data.get('name', '').strip()
    qty = int(data.get('quantity', 0))
    unit = data.get('unit', '罐').strip()
    mfg_date = format_date_str(data.get('mfg_date', ''))
    exp_date = format_date_str(data.get('exp_date', ''))
    in_date = format_date_str(data.get('in_date', ''))
    staff = data.get('staff', '').strip()

    if not name:
        return jsonify({"error": "請輸入品名"}), 400

    with DBConn() as db:
        db.execute("""
            INSERT INTO inventory (name, initial_qty, current_qty, unit, mfg_date, exp_date, in_date, staff, status)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'active')
        """, (name, qty, qty, unit, mfg_date, exp_date, in_date, staff))
    return jsonify({"success": True})

@app.route('/api/items/<int:item_id>', methods=['PUT'])
def update_item(item_id):
    data = request.json or {}
    with DBConn() as db:
        db.execute("""
            UPDATE inventory
            SET name = ?, initial_qty = ?, current_qty = ?, unit = ?, mfg_date = ?, exp_date = ?, in_date = ?, staff = ?, notes = ?
            WHERE id = ?
        """, (
            data.get('name', '').strip(),
            int(data.get('initial_qty', 0)),
            int(data.get('current_qty', 0)),
            data.get('unit', '罐'),
            format_date_str(data.get('mfg_date', '')),
            format_date_str(data.get('exp_date', '')),
            format_date_str(data.get('in_date', '')),
            data.get('staff', '').strip(),
            data.get('notes', ''),
            item_id
        ))
    return jsonify({"success": True})

@app.route('/api/audit/<int:item_id>', methods=['POST'])
def audit_item(item_id):
    try:
        data = request.json or {}
        actual_qty = int(data.get('actual_qty', 0))
        auditor = data.get('auditor', '').strip()
        audit_date = format_date_str(data.get('audit_date', datetime.now().strftime('%Y-%m-%d')))
        notes = data.get('notes', '')

        with DBConn() as db:
            cursor = db.execute("SELECT * FROM inventory WHERE id = ?", (item_id,))
            item = cursor.fetchone()
            if not item:
                return jsonify({"error": "找不到品項"}), 404

            item_dict = dict(item)
            status = 'completed' if actual_qty == 0 else 'active'
            
            db.execute("""
                UPDATE inventory 
                SET current_qty = ?, last_audit_date = ?, auditor = ?, notes = ?, status = ?
                WHERE id = ?
            """, (actual_qty, audit_date, auditor, notes, status, item_id))

            db.execute("""
                INSERT INTO audit_logs (item_id, name, initial_qty, in_date, audit_date, actual_qty, auditor, notes, action_type)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                item_id, item_dict.get('name', ''), item_dict.get('initial_qty', 0),
                format_date_str(item_dict.get('in_date', '')),
                audit_date, actual_qty, auditor, notes, 'completed' if actual_qty == 0 else 'audit'
            ))

        return jsonify({"success": True})
    except Exception as e:
        print("Audit error:", e)
        return jsonify({"error": str(e)}), 500

@app.route('/api/items/<int:item_id>', methods=['DELETE'])
def delete_item(item_id):
    with DBConn() as db:
        db.execute("DELETE FROM inventory WHERE id = ?", (item_id,))
    return jsonify({"success": True})

@app.route('/api/recognize', methods=['POST'])
def recognize_image():
    if 'image' not in request.files:
        return jsonify({"error": "未提供圖片檔案"}), 400
    file = request.files['image']
    image_bytes = file.read()
    
    api_key = os.environ.get("GEMINI_API_KEY", "")
    if not api_key or not genai:
        return jsonify({"name": "辨識樣品", "mfg_date": "", "exp_date": ""})

    try:
        client = genai.Client(api_key=api_key)
        prompt = (
            "請分析這張食品包裝圖片，提取繁體中文品名與日期：\n"
            "1. 食品完整品名 (例如：桂格大燕麥片、愛之味脆瓜)\n"
            "2. 製造日期 (格式 YYYY-MM-DD，若無請留空)\n"
            "3. 有效日期 (格式 YYYY-MM-DD，若無請留空)\n"
            "請僅輸出 JSON：{\"name\": \"品名\", \"mfg_date\": \"YYYY-MM-DD\", \"exp_date\": \"YYYY-MM-DD\"}"
        )
        response = client.models.generate_content(
            model='gemini-2.5-flash',
            contents=[types.Part.from_bytes(data=image_bytes, mime_type=file.mimetype or "image/jpeg"), prompt]
        )
        text = response.text or ""
        match = re.search(r'\{.*\}', text, re.DOTALL)
        if match:
            import json
            res_json = json.loads(match.group(0))
            res_json['mfg_date'] = format_date_str(res_json.get('mfg_date', ''))
            res_json['exp_date'] = format_date_str(res_json.get('exp_date', ''))
            return jsonify(res_json)
        return jsonify({"name": text.strip()[:30], "mfg_date": "", "exp_date": ""})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/export')
def export_excel():
    month = request.args.get('month', datetime.now().strftime("%Y-%m"))
    with DBConn() as db:
        cursor = db.execute("SELECT * FROM inventory ORDER BY id ASC")
        items = cursor.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "食品存放盤點表"

    title_font = Font(name="微軟正黑體", size=14, bold=True)
    org_font = Font(name="微軟正黑體", size=10, color="555555")
    header_font = Font(name="微軟正黑體", size=10, bold=True)
    body_font = Font(name="微軟正黑體", size=10)

    header_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='B0B0B0'), right=Side(style='thin', color='B0B0B0'),
        top=Side(style='thin', color='B0B0B0'), bottom=Side(style='thin', color='B0B0B0')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    ws.merge_cells("A1:K1")
    ws["A1"] = f"食品物資存放盤點表 ({month})"
    ws["A1"].font = title_font
    ws["A1"].alignment = center_align

    ws.merge_cells("A2:K2")
    ws["A2"] = "財團法人私立天主教中華聖母社會福利慈善事業基金會 附設嘉義縣私立隆興社區長照機構(團體家屋)"
    ws["A2"].font = org_font
    ws["A2"].alignment = center_align

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 26

    headers = ["序號", "食品品名", "入庫數", "製造日期", "有效日期", "入庫日期", "入庫人員", "最近盤點日", "目前庫存", "盤點人", "說明備註"]
    ws.append(headers)

    for col in range(1, 12):
        c = ws.cell(row=3, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center_align
        c.border = thin_border

    for idx, row in enumerate(items, 1):
        d = dict(row)
        mfg = format_date_str(d.get('mfg_date')) or '-'
        exp = format_date_str(d.get('exp_date')) or '-'
        in_d = format_date_str(d.get('in_date')) or '-'
        last_audit = format_date_str(d.get('last_audit_date')) or '-'
        
        init_q = d.get('initial_qty', '-')
        cur_q = d.get('current_qty', '-')
        unit = d.get('unit', '')

        row_data = [
            idx, d.get('name', ''), f"{init_q} {unit}".strip(), mfg, exp, in_d,
            d.get('staff', '') or '-', last_audit, f"{cur_q} {unit}".strip(),
            d.get('auditor', '') or '-', d.get('notes', '') or '-'
        ]
        ws.append(row_data)

        curr_r = ws.max_row
        ws.row_dimensions[curr_r].height = 22
        for col in range(1, 12):
            cell = ws.cell(row=curr_r, column=col)
            cell.font = body_font
            cell.border = thin_border
            cell.alignment = left_align if col in [2, 11] else center_align

    widths = {1: 6, 2: 24, 3: 10, 4: 14, 5: 14, 6: 14, 7: 12, 8: 14, 9: 12, 10: 12, 11: 18}
    for col, w in widths.items():
        ws.column_dimensions[chr(64 + col)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"食品物資存放盤點表_{month}.xlsx"
    )

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=True)
